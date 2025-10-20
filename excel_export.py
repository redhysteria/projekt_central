from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from models import Quote, QuoteItem, MonthlyDistribution
from business_logic import BusinessLogic

class ExcelExporter:
    def __init__(self):
        self.business_logic = BusinessLogic()
    
    def export_quote(self, quote_id, filepath):
        """Export quote to Excel file with formatting matching Google Sheets structure"""
        quote = Quote.query.get(quote_id)
        if not quote:
            raise ValueError(f"Quote with id {quote_id} not found")
        
        items = QuoteItem.query.filter_by(quote_id=quote_id).all()
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Wycena"
        
        # Define styles
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers (columns A-W)
        headers = [
            'Zadanie', 'KTO/CO', 'Miesiąc realizacji (zalecany)', 
            'Zalecana liczba godzin pracy specjalisty (przybliżona) ALBO tys. znaków tekstu copy [j.m.]',
            'Cena za szt.', 'Cena razem', 'J.m. - klient', 'Cena na projekt - klient',
            'Miesiąc realizacji - klient', 'Kwota LB', 'Zadanie'
        ]
        
        # Add month headers (L-W)
        for month in range(1, 13):
            headers.append(f'Miesiąc {month:02d}')
        
        # Add header row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        # Add data rows
        row = 2
        for item in items:
            # Get monthly distribution for this item
            monthly_dist = {}
            for dist in MonthlyDistribution.query.filter_by(quote_item_id=item.id):
                monthly_dist[dist.month_number] = dist.amount
            
            # Row data
            row_data = [
                item.task_name,  # A - Zadanie
                item.specialist_type,  # B - KTO/CO
                item.month_execution,  # C - Miesiąc realizacji (zalecany)
                item.hours_or_units,  # D - Zalecana liczba godzin/znaków
                item.price_per_unit,  # E - Cena za szt.
                item.total_price,  # F - Cena razem
                item.client_units,  # G - J.m. - klient
                item.client_price,  # H - Cena na projekt - klient
                item.client_month,  # I - Miesiąc realizacji - klient
                item.client_price if 'LB' in item.task_name else '',  # J - Kwota LB
                item.task_name  # K - Zadanie (duplicate)
            ]
            
            # Add monthly distribution (L-W)
            for month in range(1, 13):
                row_data.append(monthly_dist.get(month, 0))
            
            # Write row data
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col >= 12:  # Monthly columns
                    cell.alignment = center_alignment
            
            row += 1
        
        # Add totals row
        totals_row = row
        monthly_totals = self.business_logic.calculate_monthly_totals(quote_id)
        
        # Calculate totals
        total_client_price = sum(item.client_price for item in items)
        
        # Totals row data
        totals_data = [
            'SUMA', '', '', '', '', '', '', total_client_price, '', '', ''
        ]
        
        # Add monthly totals
        for month in range(1, 13):
            totals_data.append(monthly_totals[month])
        
        # Write totals row
        for col, value in enumerate(totals_data, 1):
            cell = ws.cell(row=totals_row, column=col, value=value)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            if col >= 12:
                cell.alignment = center_alignment
        
        # Auto-adjust column widths
        column_widths = {
            'A': 30,  # Zadanie
            'B': 25,  # KTO/CO
            'C': 35,  # Miesiąc realizacji
            'D': 50,  # Zalecana liczba godzin
            'E': 12,  # Cena za szt.
            'F': 12,  # Cena razem
            'G': 12,  # J.m. - klient
            'H': 20,  # Cena na projekt - klient
            'I': 25,  # Miesiąc realizacji - klient
            'J': 12,  # Kwota LB
            'K': 30   # Zadanie (duplicate)
        }
        
        # Set column widths
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Set monthly column widths (L-W)
        for month in range(12, 24):  # Columns L to W
            ws.column_dimensions[get_column_letter(month)].width = 15
        
        # Add quote info at the top
        info_row = 1
        ws.insert_rows(1, 3)  # Insert 3 rows at the top
        
        ws.cell(row=1, column=1, value=f"Wycena: {quote.name}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        ws.cell(row=2, column=1, value=f"Data utworzenia: {quote.created_at.strftime('%Y-%m-%d %H:%M')}")
        ws.cell(row=3, column=1, value=f"Data aktualizacji: {quote.updated_at.strftime('%Y-%m-%d %H:%M')}")
        
        # Save workbook
        wb.save(filepath)
        
        return filepath
